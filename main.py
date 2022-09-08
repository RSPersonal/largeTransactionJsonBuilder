import openpyxl
import json
from progress.bar import Bar

if __name__ == '__main__':
    input_user = input("Please enter the filename for the target sheet. *With .xlsx extension: ")
    dataframe = openpyxl.load_workbook(input_user)
    dataframe1 = dataframe.active
    external_values = []
    for row in dataframe1.iter_rows(2, 2):
        for cell in row:
            external_values.append(cell.value)

    internal_values = []
    for row in dataframe1.iter_rows(1, 1):
        for cell in row:
            internal_values.append(cell.value)

    index_for_columns = 0
    index_for_rows = 2
    column_for_json_object = (dataframe1.max_column + 1)
    max_column = dataframe1.max_column

    progress_bar = Bar('Processing', max=(dataframe1.max_row - 1))
    for row in dataframe1.iter_rows(2, dataframe1.max_row):
        row_property = []
        for cell in row:
            test = {
                "value": "" if cell.value is None else cell.value,
                "externalValue": external_values[index_for_columns],
                "internalValue": internal_values[index_for_columns]
            }
            row_property.append(test)
            if index_for_columns == (max_column - 1):
                index_for_columns = 0
            else:
                index_for_columns += 1

        if index_for_rows > 2:
            dataframe1.cell(index_for_rows,
                            column_for_json_object).value = json.dumps(row_property, default=str)
        index_for_rows += 1
        progress_bar.next()
    dataframe.save('newExcelSheetWithJsonObject.xlsx')
    progress_bar.finish()
    print("New sheet is saved with filename: newExcelSheetWithJsonObject.xlsx")
